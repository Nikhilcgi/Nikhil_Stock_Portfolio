from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

import openpyxl

from portfolio_tracker.brokers.angelone import classify_dp_description, parse_tradebook


class AngelOneTests(unittest.TestCase):
    def test_execution_fills_use_trade_id_not_order_id(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "angel.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "TradesAndCharges"
            sheet.cell(7, 1, "2026-03-02 00:00:00.0")
            sheet.cell(7, 2, "2026-08-27 23:59:59.0")
            headers = [
                "Scrip/Contract",
                "Buy/Sell",
                "Buy Price",
                "Sell Price",
                "Quantity",
                "Brokerage",
                "GST",
                "STT",
                "Sebi Tax",
                "Exchange Turnover Charges",
                "Stamp Duty",
                "Other Charges",
                "IPFT Charges",
                "Order Type",
                "Segment",
                "Exchange",
                "Order ID",
                "Trade ID",
                "Date",
            ]
            for column, value in enumerate(headers, start=1):
                sheet.cell(35, column, value)
            base = ["Example Ltd", "Buy", 100, None, 2, 1, 0.18, 0.2, 0.01, 0.04, 0.03, 0, 0, "Delivery", "CAPITAL", "NSE", "ORDER1", "TRADE1", date(2026, 4, 1)]
            for column, value in enumerate(base, start=1):
                sheet.cell(36, column, value)
            base[17] = "TRADE2"
            base[4] = 3
            for column, value in enumerate(base, start=1):
                sheet.cell(37, column, value)
            workbook.save(path)

            parsed = parse_tradebook(path, "angel-test")

        self.assertEqual(parsed.period.start, date(2026, 3, 2))
        self.assertEqual(len(parsed.trades), 2)
        self.assertEqual(parsed.trades["trade_uid"].nunique(), 2)
        self.assertEqual(parsed.trades["broker_order_id"].nunique(), 1)
        self.assertEqual(set(parsed.trades["tax_lot_quality"]), {"EXACT_FILL"})

    def test_dp_descriptions_remain_semantically_distinct(self):
        self.assertEqual(classify_dp_description("EP-CR reference"), "EARLY_PAYIN_CREDIT")
        self.assertEqual(classify_dp_description("EP-DR reference"), "EARLY_PAYIN_DEBIT")
        self.assertEqual(
            classify_dp_description("AUTO CA-AUTO CA CURRENT BALANCE CREDIT"),
            "AUTO_CORPORATE_ACTION_CREDIT",
        )
        self.assertEqual(classify_dp_description("INTDEP-CR transfer"), "INTERDEPOSITORY_CREDIT")


if __name__ == "__main__":
    unittest.main()
